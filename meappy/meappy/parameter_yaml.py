import yaml
from os import path
from pathlib import Path
from datetime import datetime
import openpyxl
import pathlib
import re, string
import sys, argparse
from copy import deepcopy
import logging

from meappy.meappy_data import get_test_data_path
from meappy.configuration import (USER_PATHS, USER, XLCOL, XL_COLS, LOG_FILE,
                                  COMPOSITE_ROW_ID)


logging.basicConfig(format='%(levelname)s: %(asctime)s >> %(message)s', 
                    datefmt='%Y-%m-%d %I:%M:%S', 
                    filename= LOG_FILE, 
                    encoding='utf-8', level=logging.DEBUG)
    

DESCRIPTION = """Create data folder and metadata from the experiment log

To use

    cd path/to/MED64_Data
    # edit MED64_ExperimentsForAnalysis.xls

    python /path/to/meap/meappy/meappy/parameter_yaml.py

Note that to configure setup, edit `path/to/MED64_Data/configuration.py`

"""

def parse_arguments(argv):
    parser = argparse.ArgumentParser(description=DESCRIPTION)
    parser.add_argument(
        "--data_root", type=str, action="store", dest="data_root",
        help="""Select user paths configuration""", default=".")
    
    parser.add_argument(
        "--slice_template_path", type=str, action="store", dest="slice_template_path",
        help="""Path to slice_template.yaml to use as a template for slice paramters""",
        default=USER_PATHS[USER]["slice_template_path"])

    parser.add_argument(
        "--protocol_template_path", type=str, action="store", dest="protocol_template_path",
        help="""Path to slice_template.yaml to use as a template for protocol paramters""",
        default=USER_PATHS[USER]["protocol_template_path"])
    
    parser.add_argument(
        "--protocol_id", type=str, action="store", dest="protocol_id",
        help="""Select protocol (i.e. a tab in the MED64_ExperimentsForAnalysis.xls)""",
        default = None)

    parser.add_argument(
        "--fields", nargs="+", type=str, action="store", dest="XL_COLS",
        help="""Select meta data fieilds for each recording (i.e. columns in a tab in the MED64_ExperimentForAnalysis.xls) (Default: XL_COLS in configuration.py)""",
        default = USER_PATHS[USER]["sample_fields"])

    parser.add_argument(
        "--sample_ids", nargs="+", type=str, action="store", dest="sample_ids",
        help="""Select recording ids (i.e. rows in tab in the MED64_ExperimentsForAnalysis.xls), default to all""",
        default = None)

    args = parser.parse_args()
    return args


def set_output_paths(user=USER, user_paths=USER_PATHS, protocol_dir=r""):
    # write check for os paths
    output_path = path.join(user_paths[user]["output"], r"experiment")
    user_paths[user]["protocol_output"] = path.join(output_path, protocol_dir)


def set_input_paths(args, user=USER, user_paths=USER_PATHS):
    user_paths[user]["slice_template"] = path.join(
        user_paths[user]["meap"], args.slice_template_path
    )
    user_paths[user]["protocol_template"] = path.join(
        user_paths[user]["meap"], args.protocol_template_path
    )


def set_user(user_name, args, user_paths=USER_PATHS):
    if user_name not in user_paths:
        raise KeyError(f"User {user_name} not in known USER_PATHS {USER_PATHS.keys()}")
    set_input_paths(args, user_name)
    set_output_paths(user_name)
    return user_name


def get_date_param(date_str):
    date_dt = datetime.strptime(date_str, "%Y%m%d")
    date_param = datetime.strftime(date_dt, "%Y-%m-%d")
    return date_param


def get_time_param(time_str):
    time_dt = datetime.strptime(time_str, "%Hh%Mm%Ss")
    time_param = datetime.strftime(time_dt, "%I:%M:%S %p")
    return time_param


def xl_to_protocol_params(researchers, protocol_name, protocol_param_template):
    """Most of these parameters are manually set in the protocol_parameters.yaml"""
    protocol_params = protocol_param_template
    protocol_params["protocol_metadata"]["researchers"] = researchers
    protocol_params["protocol_metadata"]["protocol"] = protocol_name
    return protocol_params


def mkdir_slice(slice_id=""):
    protocol_dir_path = pathlib.Path(USER_PATHS[USER]["protocol_output"])
    slice_dir_path = protocol_dir_path / slice_id
    slice_dir_path.mkdir(mode=511, parents=True, exist_ok=True)


def create_protocol_params(researchers, protocol_name):
    with open(USER_PATHS[USER]["protocol_template"], "r") as file:
        protocol_param_template = yaml.safe_load(file)
    protocol_params = xl_to_protocol_params(researchers, protocol_name, protocol_param_template)
    protocol_params_dump = yaml.dump(
        protocol_params, sort_keys=False, indent=4, default_flow_style=False
    )
    protocol_dump_file = (
        pathlib.Path(USER_PATHS[USER]["protocol_output"]) / r"protocol_parameters.yaml"
    )
    mkdir_slice()
    with open(protocol_dump_file, "w", encoding="utf-8") as yaml_file:
        yaml_file.write(protocol_params_dump)
    logging.info(f"Parameter protocol written to: {protocol_dump_file}")


def mock_excel():
    """
    Deprecated
    
    Create mock excel row data for dev and testing.
    Alternative code is for including widgets in Jupyter Notebooks
    """
    example_excel_row = (
        "20220211\tEVD\tEVD\tVTA\t\tDose/Response\t"
        "EM2 (1pM,10pM,100pM,1nM,10nM,100nM); CTAP (1uM)\tZadina\t15h24m23s\tY\t"
        "Chs:28,33,34,35,36,37,41,42,43,47,46,50,51,54. Maybe:27,26,25,44,52,45\t"
        "Y\tY\tGabazine in all solutions. Started, ran until 1pM EM2, stopped. Washed out and restarted form baseline (computer was lacking memory)"
    )

    xl_row = example_excel_row
    # xl_row = widgets.Text(value=example_excel_row, disabled=False, layout=ext_xlrow, display='flex',
    #     flex_flow='column', flex_wrap='wrap')

    xl_cols = [
        "date",
        "cut_by",
        "run_by",
        "region",
        "project",
        "experiment_type",
        "drugs_dose_used",
        "vendor_batch",
        "slice_time",
        "is_photo_saved",
        "notes_issues",
        "is_exported",
        "is_sorted",
        "notes",
    ]  # XL_COLS
    xl = dict(zip(xl_cols, xl_row.split("\t")))
    # pprint(xl.items())
    return xl



def clean_xl_col_names(sheet_data):
    """
    Gets and cleans the None values from an excel sheet object read by the openpyxl module.
    Args: 
        sheet_data: openpyxl object, tuple of worksheet cells from excel notebook
    Returns:
        values: list, values from worksheet data with all None values stripped from
        the terminal end.
    """
    values = [cell.value for cell in sheet_data]
    while values[-1] is None:
        values = values[:-1]
    return values


def check_xl_col_names(col_names, raw_columns):
    """Checks a list of excel workbook column names against the raw columns 
    read from a new worksheet. The raw_column names may contain extra None
    values at the end.
    """
    if col_names != clean_xl_col_names(raw_columns):
        raise ValueError(f'Columns of Experiment Spreadsheet tab must be \
        identical to other tabs.')
    return None


def filter_protocol_ids(protocol_ids):
    """
    how to create a generator from a generator?
    """
    print(protocol_ids)
    return True


def read_xl_wb_indices(wb):
    """
    Column names will be checked against each other tab to verify column name 
    consistency. Will throw error if unexpected column names. 
    Skips the "Index" tab.
    """
    col_names = None
    
    for tab in wb.sheetnames:
        if tab == 'Index':
            continue
        
        sheet_data = wb[tab]
        new_col_names = sheet_data['1']
        if col_names is None:
            col_names = clean_xl_col_names(new_col_names)
        check_xl_col_names(col_names, new_col_names)
    return col_names


def verify_row_id_elements(ws):
    """Find and verify that columns used to create composite row_id are valid.
    """
    col_names = ws['1']
    row_id_elements = []
    for i, name in enumerate(col_names):
        if name.value == COMPOSITE_ROW_ID[0]:
            row_id_elements.append(i+1)  # (i+1) because openpyxl indexing starts at 1 not 0
        if name.value == COMPOSITE_ROW_ID[1]:
            row_id_elements.append(i+1)
    if len(row_id_elements) < 2:
        raise ValueError("Row_ID must be composed of two elements")
    return row_id_elements
            
    
def read_xl_sheet_rows(ws, col_names):
    """
    Reads rows in an excel workbook sheet.
    Args:
        ws: openpyxl worksheet, containing experiment data from one
        sheet of a workbook
        col_names: names of columns in worksheet to retrieve
    Return:
        rows: dict, values in rows returned
    """
    num_rows = ws.max_row
    num_cols = len(col_names)  # ws.max_column
    row_dict = dict()
    
    row_id_elements = verify_row_id_elements(ws)
    for r in range(2, num_rows):
        row_id = str(ws.cell(row=r, column=row_id_elements[0]).value) + "_" + \
                str(ws.cell(row=r, column=row_id_elements[1]).value)
        row_data = ws[str(r)]
        row_values = [cell.value for cell in row_data]
        if row_values[0] is not None:
            row_dict[row_id] = {alphanum(k):v.value for k,v in zip(col_names, row_data)}
    return row_dict


def get_xls_data(xls_file):
    """
    Gets the index data of an excel spreadsheet to use. 
    Args:
        xls_file: file path, of xlsx file with experiment protocols in tabs 
        and slice metadata in each row.
    Returns:
        dict, tab names keys with values as dicts with keys of row_IDs 
        corresponding to slice id and values of 
    """
    # Open Workbook
    wb = openpyxl.load_workbook(filename=xls_file, data_only=True)
    col_names = read_xl_wb_indices(wb)
    tab_dict = dict()
    for tab in wb.sheetnames: 
        if tab == 'Index':
            continue
        row_dict = read_xl_sheet_rows(wb[tab], col_names)
        tab_dict[tab] = row_dict    
    return col_names, tab_dict
    

def alphanum(str):
    """
    Arg: string
    Returns: string, composed of only alphanumeric and underscore. This is used to 
    create code identifiers and filenames from human readable strings by removing 
    special characters and spaces.
    """
    pattern = re.compile('[\W]+') 
    return pattern.sub('', str.lower().replace("/", "_")
                       .replace("#", "").strip().replace(" ", "_"))
    
    
def clean_identifiers(name_list):
    """Takes a list of strings and returns a list of strings composed of only 
    alphanumeric and underscore. This is used to create code identifiers and 
    filenames from human readable strings by removing special characters and spaces.
    """
    clean_list = [alphanum(str) for str in name_list]
    return clean_list
    
        
def read_xl_row(row_name, tab_name, xl, slice_params):
    date_str = str(xl[XLCOL["date"]]).strip()
    time_str = xl[XLCOL["slice_time"]]

    try:
        slice_id = date_str + "_" + time_str
        slice_params["slice_metadata"]["date"] = get_date_param(date_str)
        slice_params["slice_metadata"]["time"] = get_time_param(time_str)
    except:
        raise ValueError(f'Date/Time format error in ' + \
                         f'time_str{time_str}, date_str: {date_str}')
        
    slice_params["slice_metadata"]["notes"] = (
        "Channels: " + str(xl[XLCOL["notes_issues"]]) + \
        "\nMisc: " + str(xl[XLCOL["notes"]])
    )
    slice_params["slice_metadata"]["region"] = xl[XLCOL["region"]]
    slice_params["slice_metadata"]["type"] = xl[XLCOL["experiment_type"]]
    slice_params["slice_metadata"]["protocol"] = tab_name
    slice_params["slice_metadata"]["cut_by"] = xl[XLCOL["cut_by"]]
    slice_params["slice_metadata"]["run_by"] = xl[XLCOL["run_by"]]
    slice_params["slice_metadata"]["recording_site"] = "[dorsal, medial, ventral]"

    unit_timestamps_filename = slice_id + "_units_ts.mat"
    treatments_filename = slice_id + "_treatments.csv"
    electrode_filename = slice_id + "_unit_electrode.csv"
    image_filename = slice_id + ".jpg"
    slice_params["paths"]["base"] = r"experiment"
    slice_params["paths"]["protocol"] = tab_name
    slice_params["paths"]["slice"] = slice_id
    slice_params["paths"]["unit"] = unit_timestamps_filename
    slice_params["paths"]["treatment"] = treatments_filename
    slice_params["paths"]["photo"] = image_filename

    return slice_params
        
    
def clean_researchers_list(protocol_dict):
    """
    Arg: dict, keys are slice_ids values are slice_parameter_template structure
    Returns: str, Cleaned up list of researchers from experiments in a protocol.
    """
    researchers = ", ".join(set([row['slice_metadata']['run_by'] for row in protocol_dict.values()]))
    return researchers
    
    
def xl_to_slice_params(xl_dict, slice_param_template):
    """
    Fill in the slice parameter yaml dict to dump to yaml file
    REPLACE xl with row
    """
    params_dict = dict()
    
    debug_list = [(n, d) for n, d in xl_dict.items()]
    
    for tab_name, tab_dict in xl_dict.items():
        params_dict[tab_name] = dict()
        for row_name, xl in tab_dict.items():
            # print(f'\nrow_name: {row_name}')
            try:
                row_params = read_xl_row(row_name, tab_name, xl, slice_param_template)
            except Exception as e:
                error_message = (f'Date/Time format error in ' + \
                                 f'tab: {tab_name}, row: {row_name}, {e}')
                logging.error(e)
                logging.error(error_message)
                continue

            # print(f'\ntab_name: {tab_name} == row_name: {row_name}')
            # print(f'row_params: {row_params}')
            params_dict[tab_name][row_name] = deepcopy(row_params)
        protocol_name = tab_name
        researchers = clean_researchers_list(params_dict[tab_name])
        
        set_output_paths(protocol_dir=protocol_name)
        create_protocol_params(researchers, protocol_name)
    # print(params_dict)
    return params_dict   
  
    
def write_slice_params(slice_params):       
    slice_params_dump = yaml.dump(
        slice_params, sort_keys=False, indent=4, default_flow_style=False
    )
    slice_dump_filepath = (
        pathlib.Path(USER_PATHS[USER]["protocol_output"])
        / slice_params["paths"]["slice"]
        / r"slice_parameters.yaml"
    )
    mkdir_slice(slice_params["paths"]["slice"])

    with open(slice_dump_filepath, "w", encoding="utf-8") as yaml_file:
        yaml_file.write(slice_params_dump)
        logging.info(f"Parameter slice params written to: {slice_dump_filepath}")
    return None
    
    
def create_slice_params(xl_dict):
    """
    xl == tab_dict; xl_dict
    
    params_dict[tab_name][row_name]
    """   
    with open(USER_PATHS[USER]["slice_template"], "r") as file:
        slice_param_template = yaml.safe_load(file)
    # print(f'slice_param_template: {slice_param_template}')
    params_dict = xl_to_slice_params(xl_dict, slice_param_template)
    
    for tab, rows in params_dict.items():
        set_output_paths(protocol_dir=tab)
        for row in rows.keys():
            slice_params = params_dict[tab][row]
            # print(f'slice_params {slice_params}')
            write_slice_params(slice_params)
    return None

    
def test_square():
    """
    testing package example
    """
    n = 2
    assert n*n == 4
    

def main(argv):
    """for development and testing, use:
    `xl = mock_excel()`
    """

    args = parse_arguments(argv)
    print(f'ARGS: {args}')
    
    user = set_user(USER, args)  # this functionality moved to parse_arguments()
    
    ## new to get all files
    xls_file = USER_PATHS[USER]["exp_xlsx"]
    col_names, tab_dict = get_xls_data(xls_file)
    # print(f'tab_dict keys: {tab_dict.keys()}')
    # print(clean_identifiers(col_names))
    
    
    
    create_slice_params(tab_dict)
        

if __name__ == "__main__":
    main(sys.argv)
